import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

from db.database import (
    get_students_by_course, get_all_subjects, get_report_by_course,
    get_report_by_student, get_summary_by_trimester, get_student,
    get_subjects_for_course, get_subjects_by_group,
)
from config import COURSE_TO_LEVEL, LEVEL_LABELS, PERIOD_LABELS

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="3a3a3a", end_color="3a3a3a", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
STUDENT_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
PERIOD_FILL = PatternFill(start_color="1e1e1e", end_color="1e1e1e", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
PROM_FILL = PatternFill(start_color="4a6cf7", end_color="4a6cf7", fill_type="solid")

TIMESTAMPS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")


def _ensure_exports_dir():
    os.makedirs(TIMESTAMPS_DIR, exist_ok=True)


def _apply_border(ws, row, col):
    ws.cell(row=row, column=col).border = THIN_BORDER


def _style_header(ws, row, col, text="", merge_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if merge_to and merge_to > col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
        ws.cell(row=row, column=col).alignment = CENTER
    return cell


def export_grades_to_excel(grade_level):
    """Export all trimesters for a course — ACTA style per student."""
    _ensure_exports_dir()
    students = get_students_by_course(grade_level)
    subjects = get_all_subjects(grade_level)
    if not subjects:
        level_key = COURSE_TO_LEVEL.get(grade_level)
        if level_key:
            subjects = get_subjects_by_group(level_key)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{grade_level}"[:31]
    periods = ["T1", "T2", "T3"]
    col = 1
    ws.cell(row=1, column=col, value="CÓDIGO")
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    col += 1
    ws.cell(row=1, column=col, value="ESTUDIANTE")
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    col += 1
    for period in periods:
        start_col = col
        for subj in subjects:
            ws.cell(row=1, column=col, value=period)
            ws.cell(row=2, column=col, value=subj["name"])
            col += 1
        end_col = col - 1
        if start_col <= end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            ws.cell(row=1, column=start_col).alignment = CENTER
        col_avg = col
        ws.cell(row=1, column=col_avg, value=period)
        ws.cell(row=2, column=col_avg, value="PROM")
        ws.cell(row=2, column=col_avg).font = Font(name="Calibri", bold=True, color="4a6cf7", size=11)
        col += 1
    for c in range(1, col):
        cell = ws.cell(row=2, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    for c in range(1, col):
        cell = ws.cell(row=1, column=c)
        cell.fill = PERIOD_FILL
        cell.font = Font(name="Calibri", bold=True, size=10, color="aaaaaa")
        cell.border = THIN_BORDER
        cell.alignment = CENTER
    report = get_report_by_course(grade_level)
    grade_map = {}
    for r in report:
        key = (r["code"], r["subject_name"])
        if key not in grade_map:
            grade_map[key] = {}
        grade_map[key][r["period"]] = r["score"]
    for row_idx, stu in enumerate(students, start=3):
        ws.cell(row=row_idx, column=1, value=stu.get("code", ""))
        ws.cell(row=row_idx, column=1).font = STUDENT_FONT
        ws.cell(row=row_idx, column=1).alignment = CENTER
        _apply_border(ws, row_idx, 1)
        ws.cell(row=row_idx, column=2, value=stu.get("nombre", ""))
        ws.cell(row=row_idx, column=2).font = STUDENT_FONT
        ws.cell(row=row_idx, column=2).alignment = LEFT
        _apply_border(ws, row_idx, 2)
        col = 3
        for period in periods:
            scores = []
            for subj in subjects:
                cell = ws.cell(row=row_idx, column=col)
                key = (stu["code"], subj["name"])
                score = grade_map.get(key, {}).get(period)
                if score is not None:
                    cell.value = float(score)
                    cell.number_format = "0.0"
                    scores.append(float(score))
                cell.alignment = CENTER
                cell.border = THIN_BORDER
                cell.font = STUDENT_FONT
                col += 1
            avg_cell = ws.cell(row=row_idx, column=col)
            if scores:
                avg_cell.value = round(sum(scores) / len(scores), 2)
                avg_cell.number_format = "0.00"
            avg_cell.alignment = CENTER
            avg_cell.border = THIN_BORDER
            avg_cell.font = Font(name="Calibri", bold=True, size=11)
            col += 1
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 35
    for c in range(3, col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "C3"
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"notas_{grade_level}_{timestamp}.xlsx"
    filepath = os.path.join(TIMESTAMPS_DIR, filename)
    wb.save(filepath)
    return filepath


def export_grades_by_course_and_trimester(course, trimester):
    """Export a single course + trimester — like ACTA per-sheet."""
    _ensure_exports_dir()
    students = get_students_by_course(course)
    level_key = COURSE_TO_LEVEL.get(course)
    subjects = get_subjects_by_group(level_key) if level_key else []
    wb = Workbook()
    ws = wb.active
    period_label = PERIOD_LABELS.get(trimester, trimester)
    ws.title = f"{course} {period_label}"[:31]
    headers = ["CÓDIGO", "ESTUDIANTE"] + [s["name"] for s in subjects] + ["PROMEDIO"]
    for c, h in enumerate(headers, 1):
        _style_header(ws, 1, c, h)
    for r, stu in enumerate(students, start=2):
        ws.cell(row=r, column=1, value=stu.get("code", "")).border = THIN_BORDER
        ws.cell(row=r, column=2, value=stu.get("nombre", "")).border = THIN_BORDER
        scores = []
        for ci, subj in enumerate(subjects, start=3):
            cell = ws.cell(row=r, column=ci)
            from db.database import get_grade
            g = get_grade(stu["id"], subj["id"], trimester)
            if g:
                cell.value = float(g["score"])
                cell.number_format = "0.0"
                scores.append(float(g["score"]))
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        avg_col = 3 + len(subjects)
        avg_cell = ws.cell(row=r, column=avg_col)
        if scores:
            avg_cell.value = round(sum(scores) / len(scores), 2)
            avg_cell.number_format = "0.00"
        avg_cell.font = Font(name="Calibri", bold=True, size=11)
        avg_cell.alignment = CENTER
        avg_cell.border = THIN_BORDER
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 35
    for c in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "C2"
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"notas_{course}_{trimester}_{timestamp}.xlsx"
    filepath = os.path.join(TIMESTAMPS_DIR, filename)
    wb.save(filepath)
    return filepath


def export_grades_by_student(student_id):
    """Full academic record by student — all subjects all trimesters."""
    _ensure_exports_dir()
    stu = get_student(student_id)
    if not stu:
        return None
    report = get_report_by_student(student_id)
    subjects_map = {}
    for r in report:
        name = r["subject_name"]
        if name not in subjects_map:
            subjects_map[name] = {"T1": None, "T2": None, "T3": None}
        subjects_map[name][r["period"]] = r["score"]
    subject_names = sorted(subjects_map.keys())
    wb = Workbook()
    ws = wb.active
    ws.title = f"{stu['nombre'][:20]}"[:31]
    r = 1
    cell = ws.cell(row=r, column=1, value=f"Expediente: {stu['nombre']} ({stu['code']})")
    cell.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    r = 2
    info_data = [
        ("Código", stu.get("code", "")),
        ("Curso", stu.get("curso", "")),
        ("Sexo", stu.get("sexo", "")),
        ("Edad", stu.get("edad", "")),
        ("Turno", stu.get("turno", "")),
    ]
    for i, (k, v) in enumerate(info_data):
        cell = ws.cell(row=r, column=1 + i * 2, value=k)
        cell.font = Font(name="Calibri", bold=True, size=10, color="aaaaaa")
        ws.cell(row=r, column=2 + i * 2, value=v).font = STUDENT_FONT
    r = 4
    headers = ["ASIGNATURA", "T1", "T2", "T3", "PROMEDIO"]
    for c, h in enumerate(headers, 1):
        _style_header(ws, r, c, h)
    for s_idx, subj_name in enumerate(subject_names):
        ri = r + 1 + s_idx
        ws.cell(row=ri, column=1, value=subj_name).font = STUDENT_FONT
        ws.cell(row=ri, column=1).border = THIN_BORDER
        scores = []
        for p_idx, period in enumerate(["T1", "T2", "T3"]):
            cell = ws.cell(row=ri, column=2 + p_idx)
            score = subjects_map[subj_name][period]
            if score is not None:
                cell.value = float(score)
                cell.number_format = "0.0"
                scores.append(float(score))
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        avg_cell = ws.cell(row=ri, column=5)
        if scores:
            avg_cell.value = round(sum(scores) / len(scores), 2)
            avg_cell.number_format = "0.00"
            avg_cell.font = Font(name="Calibri", bold=True, size=11)
        avg_cell.alignment = CENTER
        avg_cell.border = THIN_BORDER
    ws.column_dimensions["A"].width = 28
    for c in range(2, 6):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B5"
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"expediente_{stu['code']}_{timestamp}.xlsx"
    filepath = os.path.join(TIMESTAMPS_DIR, filename)
    wb.save(filepath)
    return filepath


def export_grades_summary_by_trimester(trimester):
    """Summary across ALL courses for a trimester — like ACTA RESUMEN GENERAL."""
    _ensure_exports_dir()
    period_label = PERIOD_LABELS.get(trimester, trimester)
    wb = Workbook()
    ws = wb.active
    ws.title = f"RESUMEN {period_label}"[:31]
    report = get_summary_by_trimester(trimester)
    courses_map = {}
    for r in report:
        course = r["curso"]
        if course not in courses_map:
            courses_map[course] = {}
        courses_map[course][r["subject_name"]] = r
    sorted_courses = sorted(courses_map.keys())
    headers = ["CURSO", "ASIGNATURA", "TOTAL", "APROBADOS", "SUSPENSOS", "PROMEDIO"]
    for c, h in enumerate(headers, 1):
        _style_header(ws, 1, c, h)
    row = 2
    for course in sorted_courses:
        subjects_list = sorted(courses_map[course].items(), key=lambda x: x[0])
        course_start = row
        for s_name, data in subjects_list:
            ws.cell(row=row, column=1, value=course).border = THIN_BORDER
            ws.cell(row=row, column=2, value=s_name).border = THIN_BORDER
            ws.cell(row=row, column=3, value=data["total"]).border = THIN_BORDER
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=4, value=data["aprobados"]).border = THIN_BORDER
            ws.cell(row=row, column=4).alignment = CENTER
            cell_s = ws.cell(row=row, column=5, value=data["suspensos"])
            cell_s.border = THIN_BORDER
            cell_s.alignment = CENTER
            if data["suspensos"] > data["aprobados"] and data["total"] > 0:
                cell_s.font = Font(name="Calibri", size=11, color="b05858")
            ws.cell(row=row, column=6, value=data["promedio"]).border = THIN_BORDER
            ws.cell(row=row, column=6).alignment = CENTER
            ws.cell(row=row, column=6).number_format = "0.00"
            row += 1
        if len(subjects_list) > 1:
            ws.merge_cells(start_row=course_start, start_column=1, end_row=row - 1, end_column=1)
        row += 1
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"resumen_{trimester}_{timestamp}.xlsx"
    filepath = os.path.join(TIMESTAMPS_DIR, filename)
    wb.save(filepath)
    return filepath


def export_enrollment_list(school_year_id, only_debtors=False, year_label=""):
    _ensure_exports_dir()
    from db.database import get_enrollments_by_year, get_debtors, get_school_year
    if only_debtors:
        enrollments = get_debtors(school_year_id, only_unpaid=True)
        suffix_name = "deudores"
    else:
        enrollments = get_enrollments_by_year(school_year_id)
        suffix_name = "matricula"
    year = get_school_year(school_year_id) if school_year_id else None
    label = year_label or (year["label"] if year else "")
    wb = Workbook()
    ws = wb.active
    ws.title = f"Matrícula {label}"
    headers = ["Código", "Estudiante", "Curso", "Total", "Pagado", "Deuda", "Estado"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    for r, e in enumerate(enrollments, 2):
        name = e.get('student_name', '')
        total = e.get("total_amount", 0)
        paid = e.get("paid_amount", 0)
        debt = total - paid
        status = e.get("status", "pendiente")
        ws.cell(row=r, column=1, value=e.get("student_code", "")).border = THIN_BORDER
        ws.cell(row=r, column=2, value=name).border = THIN_BORDER
        ws.cell(row=r, column=3, value=e.get("grade_level", "")).border = THIN_BORDER
        ws.cell(row=r, column=4, value=total).border = THIN_BORDER
        ws.cell(row=r, column=4).number_format = '#,##0'
        ws.cell(row=r, column=5, value=paid).border = THIN_BORDER
        ws.cell(row=r, column=5).number_format = '#,##0'
        ws.cell(row=r, column=6, value=debt).border = THIN_BORDER
        ws.cell(row=r, column=6).number_format = '#,##0'
        ws.cell(row=r, column=7, value=status).border = THIN_BORDER
    widths = [14, 35, 10, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"{suffix_name}_{label}_{timestamp}.xlsx"
    filepath = os.path.join(TIMESTAMPS_DIR, filename)
    wb.save(filepath)
    return filepath


def export_student_list(grade_level=None):
    _ensure_exports_dir()
    from db.database import get_all_students, get_students_by_level
    students = get_students_by_course(grade_level) if grade_level else get_all_students()
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    headers = ["Código", "Nombre Completo", "Sexo", "Edad", "Curso", "Turno", "Tutor", "Teléfono"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    for r, stu in enumerate(students, 2):
        ws.cell(row=r, column=1, value=stu.get("code", "")).border = THIN_BORDER
        ws.cell(row=r, column=2, value=stu.get("nombre", "")).border = THIN_BORDER
        ws.cell(row=r, column=3, value=stu.get("sexo", "")).border = THIN_BORDER
        ws.cell(row=r, column=4, value=stu.get("edad", "")).border = THIN_BORDER
        ws.cell(row=r, column=5, value=stu.get("curso", "")).border = THIN_BORDER
        ws.cell(row=r, column=6, value=stu.get("turno", "")).border = THIN_BORDER
        ws.cell(row=r, column=7, value=stu.get("tutor", "")).border = THIN_BORDER
        ws.cell(row=r, column=8, value=stu.get("phone", "")).border = THIN_BORDER
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    suffix = f"_{grade_level}" if grade_level else ""
    filename = f"estudiantes{suffix}_{timestamp}.xlsx"
    filepath = os.path.join(TIMESTAMPS_DIR, filename)
    wb.save(filepath)
    return filepath
