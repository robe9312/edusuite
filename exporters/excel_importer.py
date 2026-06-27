import os
from openpyxl import load_workbook

from db.database import diff_import, apply_import, get_student_by_code, get_subject


def _parse_filename(filepath):
    basename = os.path.basename(filepath)
    parts = basename.replace(".xlsx", "").split("_")
    return parts


def import_grades_from_excel(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = []
    errors = []

    header_row = 2
    col_map = {}
    for col in range(2, ws.max_column + 1):
        period = ws.cell(row=1, column=col).value
        subj_name = ws.cell(row=header_row, column=col).value
        if period and subj_name:
            period = str(period).strip()
            subj_name = str(subj_name).strip()
            if period in ("T1", "T2", "T3"):
                col_map[col] = (period, subj_name)

    for row in range(header_row + 1, ws.max_row + 1):
        student_name = ws.cell(row=row, column=1).value
        if not student_name:
            continue

        student_code = None
        for col, (period, subj_name) in col_map.items():
            score = ws.cell(row=row, column=col).value
            if score is None:
                continue
            try:
                score = float(score)
            except (ValueError, TypeError):
                continue
            if student_code is None:
                full_name = str(student_name).strip()
                from db.database import get_all_students
                for s in get_all_students():
                    if s.get("nombre", s.get("full_name", "")).strip() == full_name:
                        student_code = s["code"]
                        break
                if student_code is None:
                    errors.append(f"Fila {row}: '{student_name}' no encontrado en DB")
                    break

            subj = get_subject_by_name(subj_name)
            if subj is None:
                errors.append(f"Fila {row}: asignatura '{subj_name}' no encontrada")
                continue

            rows.append({
                "code": student_code,
                "subject_id": subj["id"],
                "period": period,
                "score": score,
            })

    if not rows:
        return [], [], errors

    diff = diff_import(rows)
    new_rows = [r for r in diff if r["_status"] == "new"]
    same_rows = [r for r in diff if r["_status"] == "same"]
    conflict_rows = [r for r in diff if r["_status"] == "conflict"]
    error_rows = [r for r in diff if r["_status"] == "error"]

    return diff, errors, {"new": len(new_rows), "same": len(same_rows), "conflicts": len(conflict_rows), "errors": len(error_rows)}


def get_subject_by_name(name):
    from db.database import get_all_subjects
    for s in get_all_subjects():
        if s["name"].strip().lower() == name.strip().lower():
            return s
    return None


def apply_import_result(rows):
    accepted = [r for r in rows if r.get("_status") in ("new", "conflict")]
    if accepted:
        apply_import(accepted)
    return len(accepted)
